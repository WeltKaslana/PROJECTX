# 代码说明：
'''
代码功能： 基于ChromeDriver爬取taobao（淘宝）平台商品列表数据
输入参数:  KEYWORLD --> 搜索商品“关键词”；
          pageStart --> 爬取起始页；
          pageEnd --> 爬取终止页；
输出文件：爬取商品列表数据
        'Page'        ：页码
        'Num'         ：序号
        'title'       ：商品标题
        'Price'       ：商品价格
        'Deal'        ：商品销量
        'Location'    ：地理位置
        'Shop'        ：商品
        'IsPostFree'  ：是否包邮
        'Title_URL'   ：商品详细页链接
        'Shop_URL'    ：商铺链接
        'Img_URL'     ：图片链接
'''
 
# 声明第三方库/头文件
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pyquery import PyQuery as pq
import time
import openpyxl as op               #导入Excel读写库
# 全局变量
count = 1                                   # 写入Excel商品计数
 
KEYWORD = input('输入搜索的商品关键词Keyword：')# 要搜索的商品的关键词
pageStart = int(input('输入爬取的起始页PageStart：'))# 爬取起始页
pageEnd = int(input('输入爬取的终止页PageEnd：'))# 爬取终止页
 
# 启动ChromeDriver服务
options = webdriver.ChromeOptions()
# 关闭自动测试状态显示 // 会导致浏览器报：请停用开发者模式
options.add_experimental_option("excludeSwitches", ['enable-automation'])
# 解决USB错误日志问题
options.add_experimental_option('excludeSwitches', ['enable-logging'])
# 把chrome设为selenium驱动的浏览器代理；
driver = webdriver.Chrome(options=options)
# 反爬机制
driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument",
                       {"source": """Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"""})
# 窗口最大化
driver.maximize_window()
driver.get('https://s.taobao.com/')
# wait是Selenium中的一个等待类，用于在特定条件满足之前等待一定的时间(这里是20秒)。
# 如果一直到等待时间都没满足则会捕获TimeoutException异常
wait = WebDriverWait(driver,20)


# 爬虫main函数
def Crawler_main():
    try:
        # 搜索KEYWORD
        search_goods()
        # 判断pageStart是否为第1页
        if pageStart != 1:
            turn_pageStart()
        # 爬取PageStart的商品信息
        get_goods(pageStart)
        # 从PageStart+1爬取到PageEnd
        for i in range(pageStart + 1, pageEnd+1):
            page_turning(i)
            get_goods(i)
    except Exception as exc:
        print("Crawler_main函数错误！Error：{}".format(exc))


#//*[@id="J_SearchForm"]/div/div[1]/button
# 输入“关键词”，搜索
def search_goods():
    try:
        print("正在搜索: {}".format(KEYWORD))
        # 找到搜索“输入框”
        # input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#q")))
        input = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="q"]')))
        # 找到“搜索”按钮
        # submit = wait.until(
        #     EC.element_to_be_clickable((By.CSS_SELECTOR, '#J_TSearchForm > div.search-button > button')))
        submit = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="J_SearchForm"]/div/div[1]/button')))

        input.clear()  # 先清空输入框
        # 输入框写入“关键词KeyWord”
        input.send_keys(KEYWORD)
        # 点击“搜索”按键
        submit.click()
        # 搜索商品后会再强制停止2秒，如有滑块请手动操作
        time.sleep(2)
        print("搜索完成！")
    except Exception as exc:
        print("search_goods函数错误！Error：{}".format(exc))

# 翻页函数
def page_turning(page_number):
    try:
        print("正在翻页: 第{}页".format(page_number))
        # 强制等待2秒后翻页
           # 滑动到页面底端
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        # 滑动到底部后停留3s
        time.sleep(3)
        # 找到“下一页”的按钮
        submit = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="search-content-leftWrap"]/div[2]/div[4]/div/div/button[2]')))
        submit.click()
        # 判断页数是否相等
        wait.until(EC.text_to_be_present_in_element((By.XPATH, '//*[@id="search-content-leftWrap"]/div[2]/div[4]/div/div/span[1]/em'), str(page_number)))


        print("已翻至: 第{}页".format(page_number))
    except Exception as exc:
        print("page_turning函数错误！Error：{}".format(exc))
 
 
# 翻页至第pageStar页
def turn_pageStart():
    try:
        print("正在翻转:第{}页".format(pageStart))
        # 滑动到页面底端
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        # 滑动到底部后停留3s
        time.sleep(3)
        # 找到输入“页面”的表单，输入“起始页”
        pageInput = wait.until(EC.presence_of_element_located(
            (By.XPATH, '//*[@id="search-content-leftWrap"]/div[2]/div[4]/div/div/span[3]/input')))
        pageInput.send_keys(pageStart) 
        # 找到页面跳转的“确定”按钮，并且点击
        admit = wait.until(EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="search-content-leftWrap"]/div[2]/div[4]/div/div/button[3]')))
        admit.click()
        print("已翻至:第{}页".format(pageStart))
    except Exception as exc:
        print("turn_pageStart函数错误！Error：{}".format(exc))


# 获取每一页的商品信息；
def get_goods(page):
    try:
        global count
        print(f"开始爬取第{page}页商品数据...")
        
        # 等待商品列表容器加载完成
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.content--CUnfXXxv')))
        
        # 动态等待商品加载完成，增加过滤条件
        previous_count = 0
        consecutive_stable_checks = 0
        
        while consecutive_stable_checks < 3:
            html = driver.page_source
            doc = pq(html)
            
            # 改进商品选择器，过滤广告和非商品项
            items = []
            for item in doc('div.content--CUnfXXxv > div > div').items():
                # 检查是否包含商品标题和价格元素
                if item.find('.title--qJ7Xg_90 span').text() and item.find('.innerPriceWrapper--aAJhHXD4').text():
                    items.append(item)
            
            current_count = len(items)
            
            if current_count == previous_count and current_count > 0:
                consecutive_stable_checks += 1
            else:
                consecutive_stable_checks = 0
                previous_count = current_count
            
            time.sleep(0.5)
        
        print(f"第{page}页共识别出{current_count}个有效商品")
        
        # 商品解析逻辑，增加异常处理
        success_count = 0
        fail_count = 0
        
        for item in items:
            try:
                # 提取商品信息
                title = item.find('.title--qJ7Xg_90 span').text().strip()
                
                # 价格处理，增加空值检查
                price_text = item.find('.innerPriceWrapper--aAJhHXD4').text().replace('\n', '').replace('\r', '')
                price = float(price_text) if price_text else 0.0
                
                # 销量处理，增加空值检查
                deal_text = item.find('.realSales--XZJiepmt').text()
                deal_text = deal_text.replace("万", "0000").split("人")[0].split("+")[0].strip()
                deal = int(deal_text) if deal_text else 0
                
                # 其他字段提取
                location = item.find('.procity--wlcT2xH9 span').text().strip()
                shop = item.find('.shopNameText--DmtlsDKm').text().strip()
                
                # 包邮判断
                postText = item.find('.subIconWrapper--Vl8zAdQn').text()
                is_post_free = "包邮" if "包邮" in postText else "/"
                
                # 链接提取，增加空值检查
                t_url = item.find('.doubleCardWrapperAdapt--mEcC7olq').attr('href') or ''
                shop_url = item.find('.TextAndPic--grkZAtsC a').attr('href') or ''
                img_url = item.find('.mainPicAdaptWrapper--V_ayd2hD img').attr('src') or ''
                
                # 构建商品信息字典
                product = {
                    'Page': page,
                    'Num': count-1,
                    'title': title,
                    'price': price,
                    'deal': deal,
                    'location': location,
                    'shop': shop,
                    'isPostFree': is_post_free,
                    'url': t_url,
                    'shop_url': shop_url,
                    'img_url': img_url
                }
                
                # 写入Excel
                wb.cell(row=count, column=1, value=count-1)
                wb.cell(row=count, column=2, value=title)
                wb.cell(row=count, column=3, value=price)
                wb.cell(row=count, column=4, value=deal)
                wb.cell(row=count, column=5, value=location)
                wb.cell(row=count, column=6, value=shop)
                wb.cell(row=count, column=7, value=is_post_free)
                wb.cell(row=count, column=8, value=t_url)
                wb.cell(row=count, column=9, value=shop_url)
                wb.cell(row=count, column=10, value=img_url)
                
                count += 1
                success_count += 1
                
            except Exception as e:
                fail_count += 1
                print(f"解析单个商品失败: {e}")
                continue
        
        print(f"第{page}页商品爬取完成: 成功={success_count}, 失败={fail_count}")
        
    except Exception as exc:
        print(f"爬取整页失败: {exc}")      
if __name__ == '__main__':
    # 建立Excel表格
    try:
        ws = op.Workbook()                                  # 创建Workbook
        wb = ws.create_sheet(index=0)                       # 创建worsheet
        # Excel第一行：表头
        title_list = ['Num', 'title', 'Price', 'Deal', 'Location', 'Shop', 'IsPostFree', 'Title_URL',
                      'Shop_URL', 'Img_URL']
        for i in range(0, len(title_list)):
            wb.cell(row=count, column=i + 1, value=title_list[i])
        count += 1  # 从第二行开始写爬取数据
    except Exception as exc:
        print("Excel建立失败！Error：{}".format(exc))
 
    # 开始爬取数据
    Crawler_main()
 
    # 保存Excel表格
    data = time.strftime('%Y%m%d-%H%M', time.localtime(time.time()))
    Filename = "{}_{}_FromTB.xlsx".format(KEYWORD,data)
    ws.save(filename = Filename)
    print(Filename + "存储成功~")
