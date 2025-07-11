from selenium import webdriver
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import openpyxl as op
import os
import pickle
from pyquery import PyQuery as pq
from urllib.parse import quote
from datetime import datetime

# 全局变量
count = 1  # Excel写入计数
KEYWORD = input('输入搜索的商品关键词Keyword：')
pageStart = 1
pageEnd = 1

# 对关键词进行URL编码
encoded_keyword = quote(KEYWORD, safe='')

# 京东登录相关配置（核心修改1：京东配置）
COOKIE_FILE = "jd_cookies.pkl"  # Cookie文件名修改
JD_LOGIN_URL = "https://passport.jd.com/new/login.aspx"  # 京东登录页
JD_HOME_URL = "https://www.jd.com"  # 京东首页
JD_MY_URL = "https://home.jd.com"  # 京东"我的主页"

# -------------------------- 登录与Cookie处理（核心修改2：京东登录流程） --------------------------
def get_login_cookies():
    """获取京东登录Cookie并保存"""
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    # 京东反爬较严，建议先关闭无头模式调试
    # options.add_argument("--headless")
    driver = webdriver.Chrome(options=options)
    
    # 反爬机制
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument",
                          {"source": """Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"""})
    
    # 打开京东登录页
    driver.get(JD_LOGIN_URL)
    print("请在弹出的浏览器中扫码或输入账号密码登录京东...")

    # 等待登录成功（检测"我的京东"元素，核心修改：京东登录成功标识）
    WebDriverWait(driver, 180).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "a[href*='order.jd.com']"))  # 京东"我的京东"链接
    )
    print("登录成功，获取Cookie中...")

    # 获取并保存Cookie
    cookies = driver.get_cookies()
    with open(COOKIE_FILE, "wb") as f:
        pickle.dump(cookies, f)
    print(f"Cookie已保存至 {COOKIE_FILE}")
    
    return driver

def load_valid_cookies():
    """加载并验证京东Cookie有效性"""
    if not os.path.exists(COOKIE_FILE):
        print("未找到Cookie文件，需要登录获取")
        return get_login_cookies()
    
    # 初始化浏览器
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    # options.add_argument("--headless")
    driver = webdriver.Chrome(options=options)
    
    # 反爬机制
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument",
                          {"source": """Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"""})
    
    # 先访问京东首页建立上下文
    driver.get(JD_HOME_URL)
    
    # 加载Cookie
    with open(COOKIE_FILE, "rb") as f:
        cookies = pickle.load(f)
    
    driver.delete_all_cookies()
    for cookie in cookies:
        try:
            driver.add_cookie(cookie)
        except:
            continue
    
    # 访问京东"我的主页"验证登录
    driver.get(JD_HOME_URL)
    
    try:
        # 验证登录成功（京东"我的订单"元素）
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "a[href*='order.jd.com']"))
        )
        print("Cookie有效，已登录京东")
        return driver
    except:
        print("Cookie已失效，重新登录...")
        os.remove(COOKIE_FILE)
        return get_login_cookies()

# -------------------------- 爬取逻辑（核心修改3：京东页面元素选择器） --------------------------
def Crawler_main(driver):
    """主爬取函数（京东适配）"""
    global wait # 全局driver
    wait = WebDriverWait(driver, 20)
    try:
        # 搜索商品
        search_goods()
        # 处理起始页
        if pageStart != 1:
            turn_pageStart()
        get_goods(pageStart)
        # 爬取后续页面
        for i in range(pageStart + 1, pageEnd + 1):
            page_turning(i)
            get_goods(i)
    except Exception as exc:
        print(f"Crawler_main函数错误！Error：{exc}")

def search_goods():
    """搜索商品（京东搜索框/按钮选择器）"""
    try:
        print(f"正在搜索京东商品: {KEYWORD}")
        # 京东搜索框（核心修改：选择器替换）
        input = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#key')))  # 京东搜索框ID为"key"
        # 京东搜索按钮（核心修改：选择器替换）
        submit = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#search button')))
        
        input.clear()
        input.send_keys(KEYWORD)
        submit.click()
        time.sleep(2)
        
        # 京东搜索可能打开新标签页，切换到最新标签页
        handles = driver.window_handles
        if len(handles) > 1:
            driver.switch_to.window(handles[-1])
        print("京东搜索完成！")
    except Exception as exc:
        print(f"search_goods函数错误！Error：{exc}")

def page_turning(page_number):
    """翻页至指定页码（京东翻页按钮）"""
    try:
        print(f"正在翻页至京东第{page_number}页")
        time.sleep(2)
        # 京东下一页按钮（核心修改：选择器替换）
        next_btn = wait.until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR, 'a.pn-next')  # 京东下一页按钮
        ))
        next_btn.click()
        # 验证页码（京东当前页码标识）
        wait.until(EC.text_to_be_present_in_element(
            (By.CSS_SELECTOR, 'span.p-num a.curr'),  # 京东当前页码
            str(page_number)
        ))
        print(f"已翻至京东第{page_number}页")
    except Exception as exc:
        print(f"page_turning函数错误！Error：{exc}")

def turn_pageStart():
    """跳转至起始页（京东页码输入框）"""
    try:
        print(f"正在跳转至京东第{pageStart}页")
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)
        
        # 京东页码输入框（核心修改：选择器替换）
        page_input = wait.until(EC.presence_of_element_located(
            (By.CSS_SELECTOR, 'input#jumpPage'))  # 京东页码输入框
        )
        page_input.clear()
        page_input.send_keys(pageStart)
        
        # 京东跳转确认按钮（核心修改：选择器替换）
        confirm_btn = wait.until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR, 'a.btn.btn-default'))  # 京东确认按钮
        )
        confirm_btn.click()
        print(f"已跳转至京东第{pageStart}页")
    except Exception as exc:
        print(f"turn_pageStart函数错误！Error：{exc}")

# 懒加载解决（京东商品列表懒加载较明显）
def scroll_step_by_step(scroll_distance=600, pause_time=1):
    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(0.5)
    print(f"开始滚动京东页面加载商品...")
    
    last_scroll_top = 0
    max_attempts = 50
    attempts = 0
    
    while attempts < max_attempts:
        driver.execute_script(f"window.scrollBy(0, {scroll_distance});")
        attempts += 1
        time.sleep(pause_time)
        
        current_scroll_top = driver.execute_script("return window.pageYOffset || document.documentElement.scrollTop;")
        page_height = driver.execute_script("return document.body.scrollHeight;")
        viewport_height = driver.execute_script("return window.innerHeight;")
        
        if (current_scroll_top == last_scroll_top) and (current_scroll_top + viewport_height >= page_height - 100):
            print(f"已滚动到京东页面底部（总次数：{attempts}）")
            break
        last_scroll_top = current_scroll_top
    
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(pause_time)
    print("京东商品懒加载内容已加载")

def get_goods(page):
    """获取京东商品数据（核心：京东商品元素选择器）"""
    try:
        global count
        print(f"开始爬取京东第{page}页商品数据...")
        # 等待京东商品列表加载（核心修改：容器选择器）
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '#J_goodsList')))  # 京东商品列表容器
        time.sleep(1)
        
        # 滚动加载所有商品
        scroll_step_by_step()
        
        # 解析京东商品列表（核心修改：商品项选择器）
        html = driver.page_source
        doc = pq(html)
        items = doc('li.gl-item').items()  # 京东商品项为li.gl-item
        
        success_count = 0   
        fail_count = 0
        
        for item in items:
            try:
                # 京东商品标题（核心修改：选择器）
                title = item.find('.p-name em').text().strip().replace('\n', '')
                
                # 京东商品价格（核心修改：选择器）
                price_text = item.find('.p-price .price').text().strip()
                price = float(price_text) if price_text else 0.0
                
                # 京东商品销量（核心修改：选择器）
                deal_text = item.find('.p-commit .commit').text().strip()
                deal_text = deal_text.replace("万+", "0000").replace("+", "").split("条")[0].strip()
                deal = int(float(deal_text)) if deal_text else 0
                
                # 京东商品产地（核心修改：选择器）
                location = item.find('.p-origin').text().strip().split(' ')[0]
                
                # 京东店铺名称（核心修改：选择器）
                shop = item.find('.p-shop span a').text().strip()
                
                # 京东包邮判断（核心修改：逻辑）
                is_post_free = "包邮" if "包邮" in item.find('.p-icons').text() else "/"
                
                # 商品链接（核心修改：选择器）
                t_url = item.find('.p-name a').attr('href') or ''
                if t_url.startswith('//'):
                    t_url = f'https:{t_url}'
                elif not t_url.startswith('http'):
                    t_url = f'https://item.jd.com{t_url}'
                
                # 店铺链接（核心修改：选择器）
                shop_url = item.find('.p-shop a').attr('href') or ''
                if shop_url.startswith('//'):
                    shop_url = f'https:{shop_url}'
                
                # 商品图片（核心修改：选择器）
                img_url = item.find('.p-img img').attr('src') or item.find('.p-img img').attr('data-lazy-img') or ''
                if img_url.startswith('//'):
                    img_url = f'https:{img_url}'
                
                # 写入Excel
                wb.cell(row=count, column=1, value=count-1)
                wb.cell(row=count, column=2, value=title)
                wb.cell(row=count, column=3, value=price)
                wb.cell(row=count, column=4, value=deal)
                wb.cell(row=count, column=5, value=location)
                wb.cell(row=count, column=6, value=shop)
                wb.cell(row=count, column=7, value=is_post_free)
                wb.cell(row=count, column=8, value=t_url)
                wb.cell(row=count, column=9, value=shop_url)
                wb.cell(row=count, column=10, value=img_url)
                
                count += 1
                success_count += 1
            except Exception as e:
                fail_count += 1
                continue
        
        print(f"京东第{page}页爬取完成：成功{success_count}个，失败{fail_count}个")
    
    except Exception as exc:
        print(f"get_goods函数错误！Error：{exc}")

# -------------------------- 主函数 --------------------------
if __name__ == '__main__':
    # 初始化Excel（修改文件标识为京东）
    try:
        workbook = op.Workbook()
        wb = workbook.active
        wb.title = "京东商品数据"
        title_list = ['Num', 'Title', 'Price', 'Deal', 'Location', 'Shop', 
                     'IsPostFree', 'Title_URL', 'Shop_URL', 'Img_URL']
        for i in range(len(title_list)):
            wb.cell(row=1, column=i+1, value=title_list[i])
        count = 2  # 从第2行开始写入数据
    except Exception as exc:
        print(f"Excel初始化失败！Error：{exc}")
        exit()
    
    # 获取登录后的浏览器实例
    driver = load_valid_cookies()
    
    # 执行爬取
    Crawler_main(driver)
    
    # 保存数据（修改文件名标识为京东）
    timestamp = time.strftime('%Y%m%d-%H%M', time.localtime())
    filename = f"{KEYWORD}_{timestamp}_FromJD.xlsx"  # 文件名含京东标识
    workbook.save(filename)
    print(f"数据已保存至 {filename}")
    
    # 关闭浏览器
    driver.quit()


